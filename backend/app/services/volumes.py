from __future__ import annotations

import csv
import re
from pathlib import Path
from typing import Any

from app.schemas import PileMetrics, SurveySummary


_NUMBER = re.compile(r"[-+]?\d*\.?\d+(?:[eE][-+]?\d+)?")


def parse_number(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    text = str(value).strip()
    if not text:
        return None
    match = _NUMBER.search(text.replace(",", ""))
    if not match:
        return None
    try:
        return float(match.group(0))
    except ValueError:
        return None


def patio_from_pile_id(pile_id: str) -> str | None:
    # NC_CYA_P_01 -> A ; NC_CY1_P_01 -> A (legacy numeric patio codes)
    match = re.search(r"NC_CY([ABC123])_P_\d+", pile_id.upper())
    if not match:
        return None
    token = match.group(1)
    return {"1": "A", "2": "B", "3": "C"}.get(token, token)


def normalize_pile_id(pile_id: str) -> str:
    """Map NC_CY1/2/3_* onto NC_CYA/B/C_* so report sheets can align."""
    return (
        pile_id.upper()
        .replace("NC_CY1_", "NC_CYA_")
        .replace("NC_CY2_", "NC_CYB_")
        .replace("NC_CY3_", "NC_CYC_")
    )


def _row_get(row: dict[str, str], *keys: str) -> str | None:
    lower = {k.strip().lower(): v for k, v in row.items() if k is not None}
    for key in keys:
        if key.lower() in lower:
            return lower[key.lower()]
    return None


def load_volumes_csv(path: Path) -> tuple[list[PileMetrics], SurveySummary]:
    piles: list[PileMetrics] = []
    summary = SurveySummary()

    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            if not row:
                continue
            name = (_row_get(row, "<Feature Name>", "Feature Name", "NAME", "Name") or "").strip()
            # Summary footer rows
            if not name:
                for key, value in row.items():
                    text = f"{key} {value}".lower()
                    if "net volume" in text:
                        summary.net_volume_m3 = parse_number(value) or parse_number(key)
                    if "total enclosed area" in text:
                        summary.enclosed_area_ha = parse_number(value) or parse_number(key)
                    if "total length" in text or "perimeter" in text:
                        km = parse_number(value) or parse_number(key)
                        if km is not None:
                            summary.perimeter_km = km
                continue

            if not name.upper().startswith("NC_"):
                # Skip non-pile rows
                joined = " ".join(str(v) for v in row.values() if v).lower()
                if "net volume" in joined:
                    summary.net_volume_m3 = parse_number(joined)
                continue

            metrics = PileMetrics(
                id=name,
                name=name,
                patio=patio_from_pile_id(name),
                total_volume_m3=parse_number(_row_get(row, "TOTAL_VOLUME", "Total Volume")),
                net_volume_m3=parse_number(_row_get(row, "NET_VOLUME", "Net Volume")),
                cut_volume_m3=parse_number(_row_get(row, "CUT_VOLUME", "Cut Volume")),
                fill_volume_m3=parse_number(_row_get(row, "FILL_VOLUME", "Fill Volume")),
                enclosed_area_ha=parse_number(
                    _row_get(row, "<ENCLOSED_AREA>", "ENCLOSED_AREA", "ENCLOSED_A", "Enclosed Area")
                ),
                perimeter_m=parse_number(_row_get(row, "<PERIMETER>", "PERIMETER", "Perimeter")),
                avg_elev_m=parse_number(_row_get(row, "AVG_ELEV_M")),
                min_elev_m=parse_number(_row_get(row, "MIN_ELEV_M")),
                max_elev_m=parse_number(_row_get(row, "MAX_ELEV_M")),
                properties={k: v for k, v in row.items() if v not in (None, "")},
            )
            piles.append(metrics)

    summary.pile_count = len(piles)
    summary.named_pile_count = len(piles)
    if summary.total_volume_m3 is None:
        vals = [p.total_volume_m3 for p in piles if p.total_volume_m3 is not None]
        summary.total_volume_m3 = sum(vals) if vals else None
    if summary.net_volume_m3 is None:
        vals = [p.net_volume_m3 for p in piles if p.net_volume_m3 is not None]
        summary.net_volume_m3 = sum(vals) if vals else None
    if summary.cut_volume_m3 is None:
        vals = [p.cut_volume_m3 for p in piles if p.cut_volume_m3 is not None]
        summary.cut_volume_m3 = sum(vals) if vals else None
    if summary.fill_volume_m3 is None:
        vals = [p.fill_volume_m3 for p in piles if p.fill_volume_m3 is not None]
        summary.fill_volume_m3 = sum(vals) if vals else None
    if summary.enclosed_area_ha is None:
        vals = [p.enclosed_area_ha for p in piles if p.enclosed_area_ha is not None]
        summary.enclosed_area_ha = sum(vals) if vals else None

    return piles, summary


def piles_from_geojson_features(features: list[dict]) -> list[PileMetrics]:
    piles: list[PileMetrics] = []
    for feature in features:
        props = feature.get("properties") or {}
        fid = str(props.get("feature_id") or feature.get("id") or len(piles) + 1)
        name = props.get("NAME") or props.get("Name")
        piles.append(
            PileMetrics(
                id=str(name or fid),
                name=str(name) if name else None,
                patio=patio_from_pile_id(str(name or "")),
                enclosed_area_ha=parse_number(props.get("area_ha") or props.get("ENCLOSED_A")),
                avg_elev_m=parse_number(props.get("AVG_ELEV_M")),
                min_elev_m=parse_number(props.get("MIN_ELEV_M")),
                max_elev_m=parse_number(props.get("MAX_ELEV_M")),
                centroid=(
                    [float(props["centroid_x"]), float(props["centroid_y"])]
                    if props.get("centroid_x") is not None and props.get("centroid_y") is not None
                    else None
                ),
                properties={k: v for k, v in props.items() if k not in {"centroid_x", "centroid_y"}},
            )
        )
    return piles


def merge_volume_into_piles(geo_piles: list[PileMetrics], volume_piles: list[PileMetrics]) -> list[PileMetrics]:
    by_id = {p.id: p for p in volume_piles}
    for p in volume_piles:
        by_id.setdefault(normalize_pile_id(p.id), p)
    merged: list[PileMetrics] = []
    used: set[str] = set()
    for pile in geo_piles:
        vol = by_id.get(pile.id) or by_id.get(normalize_pile_id(pile.id))
        if vol:
            used.add(vol.id)
            used.add(normalize_pile_id(vol.id))
            merged.append(
                pile.model_copy(
                    update={
                        "id": normalize_pile_id(vol.id) if vol.id.upper().startswith("NC_") else pile.id,
                        "total_volume_m3": vol.total_volume_m3,
                        "net_volume_m3": vol.net_volume_m3,
                        "cut_volume_m3": vol.cut_volume_m3,
                        "fill_volume_m3": vol.fill_volume_m3,
                        "enclosed_area_ha": vol.enclosed_area_ha or pile.enclosed_area_ha,
                        "perimeter_m": vol.perimeter_m or pile.perimeter_m,
                        "avg_elev_m": vol.avg_elev_m or pile.avg_elev_m,
                        "min_elev_m": vol.min_elev_m or pile.min_elev_m,
                        "max_elev_m": vol.max_elev_m or pile.max_elev_m,
                        "patio": vol.patio or pile.patio,
                        "name": normalize_pile_id(vol.name or vol.id)
                        if (vol.name or vol.id).upper().startswith("NC_")
                        else (vol.name or pile.name),
                    }
                )
            )
        else:
            merged.append(pile)
    for pile_id, vol in {normalize_pile_id(p.id): p for p in volume_piles}.items():
        if pile_id not in used and vol.id not in used:
            merged.append(
                vol.model_copy(
                    update={
                        "id": pile_id,
                        "name": pile_id,
                        "patio": patio_from_pile_id(pile_id),
                    }
                )
            )
    return merged


def load_volumes_xlsx(path: Path) -> tuple[list[PileMetrics], SurveySummary]:
    """Parse NCL-style calculation / final sheets (Name, Pile Name, Net Volume, Enclosed Area)."""
    from openpyxl import load_workbook

    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    header_idx = None
    for i, row in enumerate(rows[:15]):
        values = [str(v).strip().lower() if v is not None else "" for v in row]
        if any("pile name" in v for v in values) or (
            any(v == "name" for v in values) and any("net volume" in v for v in values)
        ):
            header_idx = i
            break

    piles: list[PileMetrics] = []
    summary = SurveySummary()
    if header_idx is None:
        return piles, summary

    header = [str(v).strip() if v is not None else "" for v in rows[header_idx]]

    def col(*names: str) -> int | None:
        lowered = [h.lower() for h in header]
        for name in names:
            for i, h in enumerate(lowered):
                if name in h:
                    return i
        return None

    i_patio = col("name")  # PATIO_A column often labeled Name
    i_pile = col("pile name")
    i_net = col("net volume")
    i_area = col("enclosed area")
    i_product = col("product")
    i_mass = col("mass")

    for row in rows[header_idx + 1 :]:
        if not row:
            continue
        # footer
        joined = " ".join(str(v) for v in row if v is not None)
        if "total volume" in joined.lower():
            summary.total_volume_m3 = parse_number(joined)
            summary.net_volume_m3 = summary.net_volume_m3 or summary.total_volume_m3
            continue

        pile_raw = row[i_pile] if i_pile is not None and i_pile < len(row) else None
        if not isinstance(pile_raw, str) or not pile_raw.upper().startswith("NC_"):
            continue
        pile_id = normalize_pile_id(pile_raw.strip())
        patio_raw = row[i_patio] if i_patio is not None and i_patio < len(row) else None
        patio = None
        if isinstance(patio_raw, str) and "PATIO" in patio_raw.upper():
            patio = patio_raw.upper().replace("PATIO", "").replace("_", "").replace(" ", "")[-1:]
        patio = patio or patio_from_pile_id(pile_id)

        net = parse_number(row[i_net]) if i_net is not None and i_net < len(row) else None
        area = parse_number(row[i_area]) if i_area is not None and i_area < len(row) else None
        props: dict[str, Any] = {"source": path.name}
        if i_product is not None and i_product < len(row) and row[i_product] is not None:
            props["product"] = str(row[i_product])
        if i_mass is not None and i_mass < len(row) and row[i_mass] is not None:
            props["mass_tonnes"] = parse_number(row[i_mass])

        piles.append(
            PileMetrics(
                id=pile_id,
                name=pile_id,
                patio=patio,
                net_volume_m3=net,
                total_volume_m3=net,
                enclosed_area_ha=area,
                properties=props,
            )
        )

    summary.pile_count = len(piles)
    summary.named_pile_count = len(piles)
    if summary.net_volume_m3 is None:
        vals = [p.net_volume_m3 for p in piles if p.net_volume_m3 is not None]
        summary.net_volume_m3 = sum(vals) if vals else None
        summary.total_volume_m3 = summary.net_volume_m3
    if summary.enclosed_area_ha is None:
        vals = [p.enclosed_area_ha for p in piles if p.enclosed_area_ha is not None]
        summary.enclosed_area_ha = sum(vals) if vals else None
    return piles, summary


def load_volumes(path: Path) -> tuple[list[PileMetrics], SurveySummary]:
    suffix = path.suffix.lower()
    if suffix == ".csv":
        return load_volumes_csv(path)
    if suffix in {".xlsx", ".xlsm"}:
        return load_volumes_xlsx(path)
    raise ValueError(f"Unsupported volume file type: {path.suffix}")
